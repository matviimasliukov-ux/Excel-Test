import io
import zipfile
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, numbers

st.set_page_config(page_title="Weekly Company Report Calculator", layout="wide")

# ---------- Compact sidebar styling ----------
st.markdown("""
<style>
/* Normal, compact sidebar width */
[data-testid="stSidebar"] { min-width: 18rem; max-width: 22rem; }
/* Tighten spacing inside our custom sidebar wrapper */
.sidebar-tech .block-container { padding-top: 0.5rem !important; }
.sidebar-tech .tech-item { margin: 0.25rem 0 0.5rem 0; padding-bottom: 0.25rem; border-bottom: 1px solid rgba(255,255,255,0.08); }
.sidebar-tech .tech-title { display: flex; align-items: center; justify-content: space-between; }
.sidebar-tech .tech-meta { margin: 0.1rem 0 0.25rem 0; font-size: 0.85rem; opacity: 0.85; }
/* Smaller edit button */
.sidebar-tech .stButton>button { padding: 0.2rem 0.4rem; line-height: 1; border-radius: 6px; }
/* Expander padding */
.sidebar-tech [data-testid="stExpander"] details > div { padding-top: 0.25rem; padding-bottom: 0.25rem; }
</style>
""", unsafe_allow_html=True)

# ---------- Pre-populated technicians ----------
PREPOP_TECHS = [
    {"name": "John Doe",        "rate_pct": 25.0, "truck": False, "meter": False},
    {"name": "Nathan Stevens",  "rate_pct": 25.0, "truck": False, "meter": False},
    {"name": "Spencer Monahan", "rate_pct": 25.0, "truck": False, "meter": False},
    {"name": "Mikal Segall",    "rate_pct": 25.0, "truck": False, "meter": False},
    {"name": "Jonathan Moss",   "rate_pct": 25.0, "truck": False, "meter": False},
    {"name": "Bob Rhyss",       "rate_pct": 25.0, "truck": False, "meter": False},
    {"name": "Clyde Owen",      "rate_pct": 25.0, "truck": False, "meter": False},
]

# ---------- State & helpers ----------
def init_state():
    if "employees" not in st.session_state:
        st.session_state.employees = [dict(x) for x in PREPOP_TECHS]
    if "editing_index" not in st.session_state:
        st.session_state.editing_index = None

def get_employee_by_name(name: str):
    for e in st.session_state.employees:
        if e["name"] == name:
            return e
    return None

def ensure_employee(name: str, rate_pct=25.0, truck=False, meter=False):
    if not name.strip():
        return
    e = get_employee_by_name(name.strip())
    if e is None:
        st.session_state.employees.append({
            "name": name.strip(),
            "rate_pct": float(rate_pct),
            "truck": bool(truck),
            "meter": bool(meter),
        })

def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

def format_amount_col(ws, amount_col_idx: int, bold_rows: set):
    for col_cells in ws.iter_cols(min_col=amount_col_idx, max_col=amount_col_idx, min_row=2, max_row=ws.max_row):
        for c in col_cells:
            c.alignment = Alignment(horizontal="right")
            if isinstance(c.value, (int, float)):
                c.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            if c.row in bold_rows:
                c.font = Font(bold=True)

def export_per_tech_xlsx(df_tech: pd.DataFrame, tech_info: dict, date_col: str, tech_col: str, jobfee_col: str) -> bytes:
    """
    - Amount = Job Fee * (Rate% / 100)
    - Truck ($50/day, cap $150), Meter ($25), Penguin ($6.25)
    - Charge names in far-left column, amounts in last column
    - Bold Total row; auto-size columns
    """
    d = df_tech.copy()
    d["Rate (%)"] = float(tech_info["rate_pct"])
    base = pd.to_numeric(d[jobfee_col], errors="coerce").fillna(0.0)
    d["Amount"] = base * (d["Rate (%)"] / 100.0)

    wb = Workbook()
    ws = wb.active
    ws.title = tech_info["name"][:31]

    cols = list(d.columns)
    ordered = []
    if date_col in cols: ordered.append(date_col)
    if tech_col in cols and tech_col not in ordered: ordered.append(tech_col)
    for c in cols:
        if c not in {date_col, tech_col, "Rate (%)", "Amount"}:
            ordered.append(c)
    ordered.extend(["Rate (%)", "Amount"])
    d = d[ordered]

    for r in dataframe_to_rows(d, index=False, header=True):
        ws.append(r)

    last_col_idx = ws.max_column
    amount_col_idx = last_col_idx

    bold_rows = set()
    charges_total = 0.0

    if tech_info.get("truck"):
        unique_days = pd.to_datetime(d[date_col], errors="coerce").dt.date.dropna().unique()
        truck_fee = min(3, len(unique_days)) * 50.0
        if truck_fee > 0:
            ws.append(["Truck Charge"] + [None]*(last_col_idx - 2) + [truck_fee])
            charges_total += truck_fee

    if tech_info.get("meter"):
        ws.append(["Meter Fee"] + [None]*(last_col_idx - 2) + [25.0])
        charges_total += 25.0

    ws.append(["Penguin Data Fee"] + [None]*(last_col_idx - 2) + [6.25])
    charges_total += 6.25

    data_end_row = 1 + len(d)
    data_amount_sum = 0.0
    for r in range(2, data_end_row + 1):
        val = ws.cell(row=r, column=amount_col_idx).value
        if isinstance(val, (int, float)):
            data_amount_sum += float(val)

    total = data_amount_sum + charges_total
    ws.append(["Total:"] + [None]*(last_col_idx - 2) + [total])
    total_row_idx = ws.max_row

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    for cell in ws[total_row_idx]:
        cell.font = Font(bold=True)

    bold_rows.add(total_row_idx)
    format_amount_col(ws, amount_col_idx, bold_rows)
    auto_col_width(ws)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------- UI ----------
init_state()

st.title("Weekly Company Report Calculator")

# Sidebar: one foldable bar with list + add form (clean view, inline edit per person)
with st.sidebar:
    st.markdown('<div class="sidebar-tech">', unsafe_allow_html=True)
    with st.expander("Technicians", expanded=True):

        # ---- Non-editable list with a pen button to edit one person at a time ----
        if st.session_state.employees:
            for i, e in enumerate(st.session_state.employees):
                st.markdown('<div class="tech-item">', unsafe_allow_html=True)

                c1, c2 = st.columns([8, 1])
                with c1:
                    st.markdown(f'<div class="tech-title"><b>{e["name"]}</b></div>', unsafe_allow_html=True)
                    st.markdown(
                        f'<div class="tech-meta">Rate: {e["rate_pct"]}% • Truck: {"Yes" if e.get("truck") else "No"} • Meter: {"Yes" if e.get("meter") else "No"}</div>',
                        unsafe_allow_html=True
                    )
                with c2:
                    if st.session_state.editing_index != i:
                        if st.button("✏️", key=f"edit_{i}", help="Edit"):
                            st.session_state.editing_index = i
                            st.rerun()

                # Inline editor (only for the selected tech)
                if st.session_state.editing_index == i:
                    with st.form(f"edit_form_{i}", clear_on_submit=False):
                        new_rate = st.number_input(
                            "Rate (%)", min_value=0.0, max_value=1000.0,
                            value=float(e.get("rate_pct", 25.0)), key=f"rate_{i}"
                        )
                        new_truck = st.checkbox("Truck", value=bool(e.get("truck", False)), key=f"truck_{i}")
                        new_meter = st.checkbox("Meter", value=bool(e.get("meter", False)), key=f"meter_{i}")
                        cc1, cc2 = st.columns(2)
                        save = cc1.form_submit_button("Save")
                        cancel = cc2.form_submit_button("Cancel")
                        if save:
                            e["rate_pct"] = float(new_rate)
                            e["truck"] = bool(new_truck)
                            e["meter"] = bool(new_meter)
                            st.session_state.editing_index = None
                            st.rerun()
                        if cancel:
                            st.session_state.editing_index = None
                            st.rerun()

                st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("No technicians yet.")

        # ---- Add technician (below the list, still inside the expander) ----
        st.subheader("Add technician")
        with st.form("add_tech_form", clear_on_submit=True):
            new_name = st.text_input("Name")
            new_rate_pct = st.number_input("Rate (%)", min_value=0.0, max_value=1000.0, value=25.0)
            new_truck = st.checkbox("Truck")
            new_meter = st.checkbox("Meter")
            if st.form_submit_button("Add") and new_name.strip():
                ensure_employee(new_name, rate_pct=new_rate_pct, truck=new_truck, meter=new_meter)
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# Main: upload & processing
uploaded = st.file_uploader("Upload weekly .xlsx report", type=["xlsx"])

if uploaded:
    try:
        df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Could not read Excel: {e}")
        st.stop()

    st.subheader("Preview")
    st.dataframe(df.head(20), use_container_width=True)

    cols = list(df.columns)
    if len(cols) < 3:
        st.error("The report should have at least three columns (Date, Technician, Job Fee).")
        st.stop()

    # Column mapping
    default_date = cols[0]
    guess_tech = next((c for c in ["Technician", "Tech", "Worker", "Employee", "Name"] if c in cols), cols[1])
    default_jobfee = cols[-1]

    st.markdown("#### Column mapping")
    c1, c2, c3 = st.columns(3)
    with c1:
        date_col = st.selectbox("Date column", cols, index=cols.index(default_date))
    with c2:
        tech_col = st.selectbox("Technician column", cols, index=cols.index(guess_tech))
    with c3:
        jobfee_col = st.selectbox("Job fee column (multiplied by Rate %)", cols, index=cols.index(default_jobfee))

    # Match technicians from file to our prepop list
    techs_in_file = sorted(df[tech_col].dropna().astype(str).unique())
    system_names = [e["name"] for e in st.session_state.employees]
    matched = [t for t in techs_in_file if t in system_names]

    if not matched:
        st.warning("No matching technicians between the file and the system list. Adjust names in the sidebar or the mapping.")
        st.stop()

    st.markdown("#### Technicians detected & matched")
    st.write(", ".join(matched))

    # Create ZIP of per-tech files
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, mode="w", compression=zipfile.ZIP_DEFLATED) as z:
        export_count = 0
        for t in matched:
            tech_rows = df[df[tech_col].astype(str) == t]
            if tech_rows.empty:
                continue
            info = get_employee_by_name(t)
            out_bytes = export_per_tech_xlsx(tech_rows, info, date_col, tech_col, jobfee_col)
            fname = f"{t.replace(' ', '_')}_{datetime.now().date()}.xlsx"
            z.writestr(fname, out_bytes)
            export_count += 1

    st.success(f"Prepared {export_count} file(s).")
    st.download_button(
        "Download ZIP with technician files",
        data=zbuf.getvalue(),
        file_name=f"technician_breakdowns_{datetime.now().date()}.zip",
        mime="application/zip",
    )
else:
    st.info("Upload a weekly .xlsx report to begin.")
